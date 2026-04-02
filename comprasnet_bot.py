#!/usr/bin/env python3
"""
ComprasNet Bot - Raspagem de Licitações (ComprasNet / SERPRO)
============================================================
Estratégia de captcha:
  1. Playwright (headed) -> intercepta token das requisições da página
  2. Fallback: token manual (cole do DevTools do browser)

Uso:
  python3 comprasnet_bot.py                        # Playwright auto
  python3 comprasnet_bot.py --token "P1_eyJ..."    # Token manual
  python3 comprasnet_bot.py --ano 2025             # Filtrar por ano
  python3 comprasnet_bot.py --paginas 5            # Máx. páginas de compras
"""

import argparse
import asyncio
import json
import logging
import os
import sys
import time
from datetime import datetime
from typing import Optional
from urllib.parse import quote

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ─── Configurações ────────────────────────────────────────────────────────────

BASE_API = "https://cnetmobile.estaleiro.serpro.gov.br/comprasnet-fase-externa/public/v1"
PORTAL_URL = "https://cnetmobile.estaleiro.serpro.gov.br/comprasnet-fase-externa/"

HEADERS = {
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "pt-BR,pt;q=0.9,en-US;q=0.8",
    "Referer": PORTAL_URL,
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
}

PAGE_SIZE = 50          # itens por página na listagem de compras
REQUEST_DELAY = 0.4     # segundos entre requisições (evitar throttle)


# ─── Captcha: obtém token via Playwright ──────────────────────────────────────

async def obter_token_playwright(ano: str = "2025") -> Optional[str]:
    """
    Abre o portal ComprasNet em um browser headed (visível), aguarda
    que a página faça as requisições à API e captura o token de captcha
    interceptando as requisições de rede.

    Aplica patches de stealth para evitar detecção de webdriver.
    """
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        log.warning("Playwright não instalado: pip install playwright && playwright install chromium")
        return None

    log.info("Iniciando browser para capturar token de captcha...")

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-first-run",
                "--no-default-browser-check",
                "--disable-infobars",
                "--window-size=1280,800",
                "--start-maximized",
            ],
        )

        context = await browser.new_context(
            user_agent=HEADERS["User-Agent"],
            viewport={"width": 1280, "height": 800},
            extra_http_headers={
                "Accept-Language": "pt-BR,pt;q=0.9",
                "Referer": PORTAL_URL,
            },
        )

        # Remove indicadores de webdriver
        await context.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
            Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3, 4, 5] });
            Object.defineProperty(navigator, 'languages', { get: () => ['pt-BR', 'pt', 'en'] });
            window.chrome = { runtime: {}, loadTimes: function(){}, csi: function(){} };
            Object.defineProperty(navigator, 'permissions', {
                get: () => ({ query: () => Promise.resolve({ state: 'granted' }) })
            });
        """)

        page = await context.new_page()

        captured_token = {"value": None}

        async def on_request(request):
            url = request.url
            if "/compras" in url and "captcha=" in url:
                try:
                    token = url.split("captcha=")[1].split("&")[0]
                    if token and len(token) > 100 and token.startswith("P1_"):
                        captured_token["value"] = token
                        log.info(f"Token capturado! ({token[:60]}...)")
                except Exception:
                    pass

        page.on("request", on_request)

        log.info(f"Navegando para: {PORTAL_URL}")
        await page.goto(PORTAL_URL, wait_until="domcontentloaded", timeout=60_000)

        # Aguarda até 45s pelo token; caso a página não o gere automaticamente,
        # o usuário pode navegar manualmente para a seção de licitações.
        log.info("Aguardando geração do token de captcha (máx. 45s)...")
        log.info("Se necessário, navegue até a seção de compras/licitações no browser aberto.")

        for _ in range(90):
            await asyncio.sleep(0.5)
            if captured_token["value"]:
                break

        await browser.close()

        if captured_token["value"]:
            log.info("Token obtido via Playwright com sucesso.")
        else:
            log.warning("Não foi possível capturar o token automaticamente.")

        return captured_token["value"]


def solicitar_token_manual() -> str:
    """Solicita que o usuário cole o token do captcha manualmente."""
    print("\n" + "=" * 65)
    print("COMO OBTER O TOKEN MANUALMENTE:")
    print("=" * 65)
    print("1. Abra o Chrome/Firefox e acesse:")
    print(f"   {PORTAL_URL}")
    print("2. Abra DevTools (F12) > aba Network")
    print("3. Navegue até a seção de Licitações/Compras")
    print("4. Procure uma requisição com 'captcha=P1_...' na URL")
    print("5. Copie o valor inteiro do parâmetro 'captcha'")
    print("=" * 65)
    print()
    token = input("Cole o token (P1_eyJ...): ").strip()
    if not token:
        log.error("Token não fornecido. Encerrando.")
        sys.exit(1)
    return token


# ─── Cliente da API ───────────────────────────────────────────────────────────

class ComprasNetAPI:
    def __init__(self, token: str):
        self.token = token
        self.session = requests.Session()
        self.session.headers.update(HEADERS)

    def _get(self, path: str, params: dict = None) -> Optional[dict]:
        """Faz GET na API, adicionando o token de captcha."""
        url = f"{BASE_API}{path}"
        params = params or {}
        params["captcha"] = self.token

        for tentativa in range(3):
            try:
                r = self.session.get(url, params=params, timeout=30)
                if r.status_code == 200:
                    return r.json()
                elif r.status_code == 401 or r.status_code == 403:
                    log.error(f"Token inválido/expirado (HTTP {r.status_code}). Obtenha um novo token.")
                    return None
                else:
                    log.warning(f"HTTP {r.status_code} em {url} - tentativa {tentativa + 1}")
                    time.sleep(1 * (tentativa + 1))
            except requests.RequestException as e:
                log.warning(f"Erro de rede ({tentativa + 1}/3): {e}")
                time.sleep(2 * (tentativa + 1))

        return None

    def listar_compras(self, ano: str, pagina: int = 0) -> Optional[dict]:
        filtro = json.dumps({
            "abertasParaParticipacao": False,
            "emDisputa": False,
            "emSelecaoDeFornecedores": False,
            "homologada": True,
            "deserta": False,
            "preferencialMeEpp": False,
            "modalidade": "",
            "criterioJulgamento": "",
            "unidadeCompradora": "",
            "numeroAnoCompra": ano,
        })
        return self._get("/compras", {
            "tamanhoPagina": PAGE_SIZE,
            "pagina": pagina,
            "filtro": filtro,
        })

    def detalhe_compra(self, codigo: str) -> Optional[dict]:
        time.sleep(REQUEST_DELAY)
        return self._get(f"/compras/{codigo}")

    def listar_itens(self, codigo: str, pagina: int = 0) -> Optional[dict]:
        time.sleep(REQUEST_DELAY)
        return self._get(f"/compras/{codigo}/itens", {
            "tamanhoPagina": 100,
            "pagina": pagina,
        })

    def listar_propostas(self, codigo: str, numero_item: int, pagina: int = 0) -> Optional[dict]:
        time.sleep(REQUEST_DELAY)
        return self._get(f"/compras/{codigo}/itens/{numero_item}/propostas", {
            "tamanhoPagina": 100,
            "pagina": pagina,
        })


# ─── Coleta de dados ──────────────────────────────────────────────────────────

def coletar_dados(api: ComprasNetAPI, ano: str, max_paginas: int) -> dict:
    """
    Coleta compras, itens e propostas paginando a API.
    Retorna dicionário com listas de registros.
    """
    dados = {
        "compras": [],
        "itens": [],
        "propostas": [],
    }

    # ── 1. Listar compras ──────────────────────────────────────────────────
    log.info(f"Coletando compras homologadas de {ano}...")
    total_compras = 0
    codigos_compras = []

    for pagina in range(max_paginas):
        log.info(f"  Página {pagina + 1} de compras...")
        resp = api.listar_compras(ano, pagina)

        if not resp:
            log.warning("  Resposta vazia - parando paginação.")
            break

        # A API pode retornar list ou dict com 'data'/'content'/'resultado'
        items = _extrair_lista(resp)
        if not items:
            log.info("  Sem mais resultados.")
            break

        for c in items:
            codigo = c.get("codigoCompra") or c.get("numeroCompra") or c.get("id") or ""
            registro = {
                "codigoCompra": codigo,
                "numeroCompra": c.get("numeroCompra", ""),
                "anoCompra": c.get("anoCompra", ano),
                "modalidade": _val_nested(c, "modalidade", "descricao") or c.get("modalidade", ""),
                "criterioJulgamento": _val_nested(c, "criterioJulgamento", "descricao") or "",
                "objetoCompra": c.get("objetoCompra", c.get("objeto", "")),
                "unidadeCompradora": _val_nested(c, "unidadeCompradora", "nomeUnidade") or "",
                "cnpjOrgao": _val_nested(c, "unidadeCompradora", "cnpj") or "",
                "ufOrgao": _val_nested(c, "unidadeCompradora", "uf") or "",
                "situacao": _val_nested(c, "situacao", "descricao") or c.get("situacao", ""),
                "dataPublicacao": _formatar_data(c.get("dataPublicacao", "")),
                "dataEncerramento": _formatar_data(c.get("dataEncerramentoRecebimentoPropostas", c.get("dataEncerramento", ""))),
                "valorTotalEstimado": c.get("valorTotalEstimado", ""),
                "valorTotalHomologado": c.get("valorTotalHomologado", ""),
                "linkDetalhes": f"{BASE_API}/compras/{codigo}",
            }
            dados["compras"].append(registro)
            if codigo:
                codigos_compras.append(codigo)
            total_compras += 1

        time.sleep(REQUEST_DELAY)

    log.info(f"Total de compras coletadas: {total_compras}")

    # ── 2. Itens e Propostas por compra ───────────────────────────────────
    for idx, codigo in enumerate(codigos_compras, 1):
        log.info(f"  [{idx}/{len(codigos_compras)}] Itens de compra: {codigo}")
        resp_itens = api.listar_itens(codigo)
        if not resp_itens:
            continue

        itens = _extrair_lista(resp_itens)
        for item in itens:
            num_item = item.get("numeroItem", item.get("numero", item.get("id", 0)))
            registro_item = {
                "codigoCompra": codigo,
                "numeroItem": num_item,
                "descricao": item.get("descricao", item.get("descricaoItem", "")),
                "quantidade": item.get("quantidade", ""),
                "unidadeMedida": _val_nested(item, "unidadeMedida", "descricao") or item.get("unidadeMedida", ""),
                "valorEstimado": item.get("valorEstimado", item.get("valorUnitarioEstimado", "")),
                "valorTotal": item.get("valorTotalEstimado", ""),
                "situacao": _val_nested(item, "situacao", "descricao") or item.get("situacao", ""),
                "criterioJulgamento": _val_nested(item, "criterioJulgamento", "descricao") or "",
                "beneficioMeEpp": item.get("beneficioMeEpp", ""),
            }
            dados["itens"].append(registro_item)

            # Propostas do item
            log.info(f"    Propostas item {num_item}...")
            resp_prop = api.listar_propostas(codigo, num_item)
            if resp_prop:
                propostas = _extrair_lista(resp_prop)
                for prop in propostas:
                    registro_prop = {
                        "codigoCompra": codigo,
                        "numeroItem": num_item,
                        "cnpjFornecedor": _val_nested(prop, "fornecedor", "cnpj") or prop.get("cnpj", ""),
                        "nomeFornecedor": _val_nested(prop, "fornecedor", "nomeFantasia") or _val_nested(prop, "fornecedor", "razaoSocial") or prop.get("nomeFornecedor", ""),
                        "valorProposta": prop.get("valorProposta", prop.get("valor", "")),
                        "valorNegociado": prop.get("valorNegociado", ""),
                        "classificacao": prop.get("classificacao", prop.get("colocacao", "")),
                        "situacao": _val_nested(prop, "situacao", "descricao") or prop.get("situacao", ""),
                        "dataRegistro": _formatar_data(prop.get("dataRegistroProposta", prop.get("dataRegistro", ""))),
                        "porte": _val_nested(prop, "fornecedor", "porte") or "",
                    }
                    dados["propostas"].append(registro_prop)

    return dados


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _extrair_lista(resp) -> list:
    """Extrai a lista de resultados de diferentes formatos de resposta."""
    if isinstance(resp, list):
        return resp
    if isinstance(resp, dict):
        for chave in ("data", "content", "resultado", "compras", "itens", "propostas", "items"):
            if chave in resp and isinstance(resp[chave], list):
                return resp[chave]
    return []


def _val_nested(obj: dict, *keys) -> Optional[str]:
    """Extrai valor de dicionário aninhado de forma segura."""
    for k in keys:
        if not isinstance(obj, dict):
            return None
        obj = obj.get(k)
    return obj


def _formatar_data(valor) -> str:
    """Converte timestamp ou string ISO para DD/MM/YYYY."""
    if not valor:
        return ""
    try:
        if isinstance(valor, (int, float)):
            return datetime.fromtimestamp(valor / 1000).strftime("%d/%m/%Y")
        s = str(valor)[:10]
        dt = datetime.strptime(s, "%Y-%m-%d")
        return dt.strftime("%d/%m/%Y")
    except Exception:
        return str(valor)


# ─── Exportação para Excel ────────────────────────────────────────────────────

STYLE_HEADER = {
    "font": Font(bold=True, color="FFFFFF", size=11),
    "fill": PatternFill("solid", fgColor="1F4E79"),
    "align": Alignment(horizontal="center", vertical="center", wrap_text=True),
}

STYLE_SUBHEADER = {
    "font": Font(bold=True, size=10),
    "fill": PatternFill("solid", fgColor="D6E4F0"),
    "align": Alignment(horizontal="center", vertical="center"),
}


def _escrever_cabecalho(ws, colunas: list[str]):
    ws.append(colunas)
    for col_idx, _ in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = STYLE_HEADER["font"]
        cell.fill = STYLE_HEADER["fill"]
        cell.alignment = STYLE_HEADER["align"]
    ws.row_dimensions[1].height = 30


def _ajustar_colunas(ws, min_w=10, max_w=50):
    for col in ws.columns:
        comprimento = max(
            (len(str(cell.value or "")) for cell in col),
            default=min_w,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max(comprimento + 2, min_w), max_w
        )


def exportar_excel(dados: dict, ano: str) -> str:
    """Cria o arquivo Excel com abas para Compras, Itens e Propostas."""
    nome_arquivo = f"comprasnet_{ano}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb = Workbook()

    # ── Aba Compras ────────────────────────────────────────────────────────
    ws_compras = wb.active
    ws_compras.title = "Compras"

    cols_compras = [
        "Código Compra", "Número", "Ano", "Modalidade", "Critério Julgamento",
        "Objeto da Compra", "Unidade Compradora", "CNPJ Órgão", "UF",
        "Situação", "Data Publicação", "Data Encerramento",
        "Valor Estimado (R$)", "Valor Homologado (R$)",
    ]
    _escrever_cabecalho(ws_compras, cols_compras)

    for c in dados["compras"]:
        ws_compras.append([
            c["codigoCompra"], c["numeroCompra"], c["anoCompra"],
            c["modalidade"], c["criterioJulgamento"], c["objetoCompra"],
            c["unidadeCompradora"], c["cnpjOrgao"], c["ufOrgao"],
            c["situacao"], c["dataPublicacao"], c["dataEncerramento"],
            c["valorTotalEstimado"], c["valorTotalHomologado"],
        ])

    _ajustar_colunas(ws_compras)
    ws_compras.freeze_panes = "A2"

    # ── Aba Itens ──────────────────────────────────────────────────────────
    ws_itens = wb.create_sheet("Itens")
    cols_itens = [
        "Código Compra", "Nº Item", "Descrição", "Quantidade",
        "Unidade Medida", "Valor Unitário Estimado (R$)", "Valor Total Estimado (R$)",
        "Situação", "Critério Julgamento", "Benefício ME/EPP",
    ]
    _escrever_cabecalho(ws_itens, cols_itens)

    for i in dados["itens"]:
        ws_itens.append([
            i["codigoCompra"], i["numeroItem"], i["descricao"],
            i["quantidade"], i["unidadeMedida"],
            i["valorEstimado"], i["valorTotal"],
            i["situacao"], i["criterioJulgamento"], i["beneficioMeEpp"],
        ])

    _ajustar_colunas(ws_itens)
    ws_itens.freeze_panes = "A2"

    # ── Aba Propostas ──────────────────────────────────────────────────────
    ws_prop = wb.create_sheet("Propostas")
    cols_prop = [
        "Código Compra", "Nº Item", "CNPJ Fornecedor", "Nome Fornecedor",
        "Valor Proposta (R$)", "Valor Negociado (R$)", "Classificação",
        "Situação", "Data Registro", "Porte",
    ]
    _escrever_cabecalho(ws_prop, cols_prop)

    for p in dados["propostas"]:
        ws_prop.append([
            p["codigoCompra"], p["numeroItem"], p["cnpjFornecedor"],
            p["nomeFornecedor"], p["valorProposta"], p["valorNegociado"],
            p["classificacao"], p["situacao"], p["dataRegistro"], p["porte"],
        ])

    _ajustar_colunas(ws_prop)
    ws_prop.freeze_panes = "A2"

    # ── Aba Resumo ─────────────────────────────────────────────────────────
    ws_resumo = wb.create_sheet("Resumo", 0)
    ws_resumo.column_dimensions["A"].width = 35
    ws_resumo.column_dimensions["B"].width = 20

    titulo_font = Font(bold=True, size=14, color="1F4E79")
    ws_resumo["A1"] = "RESUMO - ComprasNet Scraper"
    ws_resumo["A1"].font = titulo_font
    ws_resumo.merge_cells("A1:B1")

    resumo_dados = [
        ("", ""),
        ("Gerado em:", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Ano filtrado:", ano),
        ("Total de Compras:", len(dados["compras"])),
        ("Total de Itens:", len(dados["itens"])),
        ("Total de Propostas:", len(dados["propostas"])),
        ("", ""),
        ("Abas:", ""),
        ("  Compras", "Lista completa de compras"),
        ("  Itens", "Itens de cada compra"),
        ("  Propostas", "Propostas por item"),
    ]

    for row in resumo_dados:
        ws_resumo.append(list(row))

    wb.save(nome_arquivo)
    return nome_arquivo


# ─── Main ─────────────────────────────────────────────────────────────────────

async def main():
    parser = argparse.ArgumentParser(
        description="Bot de raspagem ComprasNet (licitações gov.br)"
    )
    parser.add_argument("--token", "-t", type=str, default="",
                        help="Token de captcha P1_eyJ... (obtido manualmente ou via env CAPTCHA_TOKEN)")
    parser.add_argument("--ano", "-a", type=str, default="2025",
                        help="Ano das compras a coletar (padrão: 2025)")
    parser.add_argument("--paginas", "-p", type=int, default=10,
                        help=f"Máximo de páginas de compras a coletar ({PAGE_SIZE} itens/pág, padrão: 10)")
    parser.add_argument("--manual", "-m", action="store_true",
                        help="Forçar entrada manual do token (sem Playwright)")
    args = parser.parse_args()

    print("\n" + "=" * 65)
    print("   COMPRASNET BOT - Raspagem de Licitações")
    print("=" * 65)

    # ── 1. Obter token ─────────────────────────────────────────────────────
    token = args.token or os.environ.get("CAPTCHA_TOKEN", "")

    if not token:
        if args.manual:
            token = solicitar_token_manual()
        else:
            log.info("Tentando capturar token via Playwright...")
            token = await obter_token_playwright(args.ano)
            if not token:
                log.warning("Playwright falhou. Solicitando token manual.")
                token = solicitar_token_manual()

    if not token:
        log.error("Token não disponível. Encerrando.")
        sys.exit(1)

    log.info(f"Token em uso: {token[:60]}...")

    # ── 2. Coletar dados ───────────────────────────────────────────────────
    api = ComprasNetAPI(token)

    log.info(f"Iniciando coleta - Ano: {args.ano} | Máx. páginas: {args.paginas}")
    inicio = time.time()

    dados = coletar_dados(api, args.ano, args.paginas)

    duracao = time.time() - inicio
    log.info(f"Coleta concluída em {duracao:.1f}s")
    log.info(f"  Compras:   {len(dados['compras'])}")
    log.info(f"  Itens:     {len(dados['itens'])}")
    log.info(f"  Propostas: {len(dados['propostas'])}")

    if not dados["compras"]:
        log.error("Nenhuma compra coletada. Verifique o token ou tente novamente.")
        sys.exit(1)

    # ── 3. Exportar ────────────────────────────────────────────────────────
    arquivo = exportar_excel(dados, args.ano)
    print(f"\n Planilha gerada: {arquivo}")
    print(f" Compras: {len(dados['compras'])} | Itens: {len(dados['itens'])} | Propostas: {len(dados['propostas'])}")
    print()


if __name__ == "__main__":
    asyncio.run(main())
